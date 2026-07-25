import sqlite3
from flask import Flask, jsonify, send_from_directory, request, send_file
from flask_cors import CORS
import datetime
import bcrypt
import jwt
from dotenv import load_dotenv
from functools import wraps
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import io

load_dotenv()
SECRET_KEY = os.getenv('SECRET_KEY')

def init_db():
    conexao = sqlite3.connect('technosensor.db')
    cursor = conexao.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT UNIQUE NOT NULL,
            nome TEXT NOT NULL,
            senha TEXT NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dados_maquina(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            status TEXT NOT NULL,
            pecas_inspecionadas INTEGER,
            pecas_aprovadas INTEGER,
            pecas_rejeitadas INTEGER,
            taxa_rejeicao REAL,
            inicio_parada TEXT,
            tempo_parado TEXT,
            media_paradas TEXT,
            status_garrafa TEXT,
            total_paradas INTEGER,
            camera TEXT,
            confianca TEXT,
            marca TEXT,
            sku TEXT,
            lote TEXT,
            envase TEXT,
            linha TEXT,
            velocidade TEXT,
            eficiencia REAL,
            disponibilidade REAL,
            opi REAL,
            oee REAL,
            mtbf TEXT,
            mttr TEXT,
            total_cameras INTEGER,
            total_triggers INTEGER,
            temperatura TEXT,
            iluminacao TEXT,
            comunicacao TEXT
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alertas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hora TEXT,
            evento TEXT,
            causa TEXT
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS falhas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            valor TEXT
        )
    ''')

    cursor.execute('SELECT COUNT(*) FROM dados_maquina')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO dados_maquina VALUES (
                1,
                'EM OPERAÇÃO',
                9842,
                9102,
                740,
                7.35,
                '14:25:47',
                '00:06:45',
                '00:07:35',
                'GARRAFA TRINCADA',
                4,
                'Câmera: CO3 - Lateral',
                'Confiança: 94%',
                'CERVEJA TIPO B',
                '600ml',
                'L150214D',
                '30/07/2026 09:15',
                'LINHA 15',
                '24.000 un/h',
                91.4,
                70.7,
                95.2,
                70.8,
                '03:45:18',
                '00:08:35',
                5,
                5,
                '32,5°C',
                'OK',
                'OK'
                )
            ''')

    cursor.execute('SELECT COUNT(*) FROM alertas')
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            'INSERT INTO alertas (hora, evento, causa) VALUES (?, ?, ?)',
            [
                ('14:10:11', 'Garrafa trincada', 'Trinca no corpo'),
                ('15:30:02', 'Garrafa trincada0', 'TRinca no ombro'),
                ('15:31:48', 'Garrafa com mancha', 'Mancha marrom'),
                ('16:31:48', 'Soda cáustica', 'Resíduo de soda'),
                ('16:32:40', 'Garrafa suja', 'Sujeira no corpo')
            ]
        )

    cursor.execute('SELECT COUNT(*) FROM falhas')
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            'INSERT INTO falhas (nome, valor) VALUES (?, ?)',
            [
                ('Trinca', '190 (43%)'),
                ('Mancha', '195 44%'),
                ('Soda cáustica', '64 (14,8%)'),
                ('Sujeira', '51 (11,8%)'),
                ('Risco', '31 (7,0%)'),
                ('Outros', '12 (8,8%)')
            ]
        )

    conexao.commit()
    conexao.close()

init_db()

def criar_hash_senha(senha):
    senha_bytes = senha.encode('utf-8')
    hash_gerado = bcrypt.hashpw(senha_bytes, bcrypt.gensalt())
    return hash_gerado.decode('utf-8')

def verificar_senha(senha_digitada, hash_salvo):
    senha_bytes = senha_digitada.encode('utf-8')
    hash_bytes = hash_salvo.encode('utf-8')
    return bcrypt.checkpw(senha_bytes, hash_bytes)

def token_obrigatorio(funcao):
    @wraps(funcao)

    def decorador(*args, **kwargs):
        token = request.headers.get('Authorization')

        if not token:
            return jsonify({'mensagem': 'Token não fornecido!'}), 401
        
        try:
            jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return jsonify({'mensagem': 'Token expirado'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'mensagem': 'Token inválida'}), 401
        
        return funcao(*args, **kwargs)
    
    return decorador

app = Flask(__name__)
CORS(app)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

@app.route('/app')
def inicio():
    return send_from_directory('../frontend', 'supervisorio.html')

@app.route('/status')
@token_obrigatorio
def status():

    conexao = sqlite3.connect('technosensor.db')
    cursor = conexao.cursor()

    cursor.execute('SELECT * FROM dados_maquina WHERE id = 1')
    linha = cursor.fetchone()

    colunas = [descricao[0] for descricao in cursor.description]
    dados = dict(zip(colunas, linha))

    cursor.execute('SELECT hora, evento, causa FROM alertas')
    dados['alertas'] = [
        {'hora': row[0], 'evento': row[1], 'causa': row[2]}
        for row in cursor.fetchall()
    ]

    cursor.execute('SELECT nome, valor FROM falhas')
    dados['top_falhas'] = [
        {'nome': row[0], 'valor': row[1]}
        for row in cursor.fetchall()
    ]

    conexao.close()
    return jsonify(dados)

    

@app.route('/cadastro', methods=['POST'])
@limiter.limit("3 per minute")
def cadastro():
    if request.method == 'OPTIONS':
        return '', 204

    dados = request.get_json(force=True)

    nome = str(dados['nome'])
    usuario = str(dados['usuario'])
    senha = str(dados['senha'])

    if not nome or not usuario or not senha:
        return jsonify({'sucesso': False, 'mensagem': 'Todos os campos são obrigatórios! '}), 400
    
    if len(senha) < 6:
        return jsonify({'sucesso': False, 'mensagem': 'A senha deve ter ao menos 6 caracteres! '}), 400
    
    senha_hash = criar_hash_senha(senha)

    try:
        conexao = sqlite3.connect('technosensor.db')
        cursor = conexao.cursor()

        cursor.execute('''
            INSERT INTO usuarios (nome, usuario, senha)
            VALUES(?, ?, ?)

        ''', (nome, usuario, senha_hash))

        conexao.commit()
        conexao.close()

        return jsonify({'sucesso': True, 'mensagem': 'Cadastro realizado com sucesso! '})
    
    except sqlite3.IntegrityError:
        return jsonify({'sucesso': False, 'mensagem': 'Usuário já existe!'}), 409

@app.route('/login', methods=['POST'])
@limiter.limit("5 per minute")
def login():
    if request.method == 'OPTIONS':
        return '', 204

    dados = request.get_json(force=True)

    usuario = dados.get('usuario')
    senha = dados.get('senha')

    if not usuario or not senha:
        return jsonify({'sucesso': False, 'mensagem': 'Usuário e senha são obrigatórios! '}), 400
    
    conexao = sqlite3.connect('technosensor.db')
    cursor = conexao.cursor()

    cursor.execute('SELECT nome, senha FROM usuarios WHERE usuario = ?', (usuario,))

    resultado = cursor.fetchone()
    conexao.close()

    if resultado is None:
        return jsonify({'sucesso': False, 'mensagem': 'Usuário ou senha incorretos'}), 401
    
    nome_encontrado, senha_hash_salva = resultado

    if not verificar_senha(senha, senha_hash_salva):
        return jsonify({'sucesso': False, 'mensagem': 'Usuário ou senha incorretos'}), 401
    
    payload = {
        'usuario': usuario,
        'nome': nome_encontrado,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=8)
    }

    token = jwt.encode(payload, SECRET_KEY, algorithm='HS256')
    
    return jsonify({'sucesso': True, 'nome': nome_encontrado, 'token': token })

@app.route('/relatorio')
@token_obrigatorio
def relatorio():
    conexao = sqlite3.connect('technosensor.db')
    cursor = conexao.cursor()

    cursor.execute('SELECT * FROM dados_maquina WHERE id = 1')
    linha = cursor.fetchone()
    colunas = [d[0] for d in cursor.description]
    dados = dict(zip(colunas, linha))

    cursor.execute('SELECT nome, valor FROM falhas LIMIT 3')
    top3_falhas = cursor.fetchall()
    conexao.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório de falhas'

    estilo_cabecalho = Font(bold=True, color='FFFFFF', size=12)
    fundo_cabecalho = PatternFill(fill_type='solid', fgColor='0F1535')
    centralizado = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:C1')
    ws['A1'] = 'Relatório das três principais falhas'
    ws['A1'].font = Font(bold=True, size=14, color='5B8CFF')
    ws['A1'].alignment = centralizado

    ws['A2'] = 'Gerado em: '
    ws['B2'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    ws.append([])

    ws.append(['STATUS DA MÁQUINA'])
    ws['A4'].font = estilo_cabecalho
    ws['A4'].fill = fundo_cabecalho

    ws.append(['STATUS', dados.get('status', '')])
    ws.append(['Peças inspecionadas', dados.get('pecas_inspecionadas', '')])
    ws.append(['Peças Aprovadas', dados.get('pecas_aprovadas', '')])
    ws.append(['Peças Rejeitadas', dados.get('pecas_rejeitadas', '')])
    ws.append(['Taxa de Rejeição', str(dados.get('taxa_rejeicao', '')) + '%'])
    ws.append(['MTBF', dados.get('mtbf', '')])
    ws.append(['MTTR', dados.get('mttr', '')])

    ws.append([])

    linha_titulo_falhas = ws.max_row + 1
    ws.append(['TOP 3 PRINCIPAIS FALHAS'])
    ws.cell(row=linha_titulo_falhas, column=1).font = estilo_cabecalho
    ws.cell(row=linha_titulo_falhas, column=1).fill = fundo_cabecalho

    ws.append(['#', 'Tipo de Falha', 'Ocorrências'])
    linha_cab = ws.max_row
    for col in range(1, 4):
        ws.cell(row=linha_cab, column=col).font = Font(bold=True)
        ws.cell(row=linha_cab, column=col).fill = PatternFill(fill_type='solid', fgColor='1E2D5A')
        ws.cell(row=linha_cab, column=col).font = Font(bold=True, color='C8D8F0')

    for i, (nome, valor) in enumerate(top3_falhas, start=1):
        ws.append([i, nome, valor])

    for col_idx, col in enumerate(ws.columns, start=1):
        largura_max = 0
        for cell in col:
            try:
                if cell.value:
                    largura_max = max(largura_max, len(str(cell.value)))
            except:
                pass
    letra = openpyxl.utils.get_column_letter(col_idx)
    ws.column_dimensions[letra].width = largura_max + 4
    
    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = f'relatorio_technosensor_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )


if __name__ == '__main__':
    app.run(debug=True)